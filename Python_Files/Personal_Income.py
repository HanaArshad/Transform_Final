# This code is automated, which mean this code will read each Excel file within the ABS_Data folder and apply the same
# code in each file because we are removing the same thing from every file


# Imports path which allows as to automate the process
from pathlib import Path

# imports PatternFills to allows us to remove colour in cells
from openpyxl.styles import PatternFill

# Identifies the path in which the Excel files are saved in this case its saved in NEW
input_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/LGA/ABS_Data'

# Import openpyxl to allow us to pick a specific sheet within the Excel workbook and load the workbook
from openpyxl import load_workbook  # pip install openpyxl

# For every Excel file in VV folder perform the following functions
for path in list(input_dir.rglob("*.xlsx*")):
    # Loads the Excel file
    wb_obj = load_workbook(filename=path)
    # Give the location of the file

    # Deletes sheets named; Contents, Dataset Info, POP, ECON, EDU, HEAL, FAM, MIG, ENV and ING_POP
    del wb_obj['Contents']
    del wb_obj['Dataset Info']
    del wb_obj['POP']
    del wb_obj['ING_POP']
    del wb_obj['ECON']
    del wb_obj['EDU']
    del wb_obj['HEAL']
    del wb_obj['FAM']
    del wb_obj['MIG']
    del wb_obj['ENV']

    # Code to delete specific items in a sheet
    # Changes in Sheet POP
    sheet = wb_obj['INC']
    # Deletes specific rows in Offenders sheet in the Excel file
    rows_to_delete = [1,2,4,5,6,9,10,11,12,15,16,17,18,21,22,23,24,27,28,29,30,33,34,35,36,37,38,39,40,41,42,43,
                      44,45,46,47,48,49,50,51,52,53,54,55,56, 57,58,59,60,61,62,63,64,65,66,67,68,70,71,72,73,74,
                      75,76,77]
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)
    # Deletes specific columns within the Excel file
    sheet.delete_cols(1, 2)
    # Remove colour in cell
    for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2000):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(fill_type="none")

    # Save the modified data into a new folder called Personal_Income
    output_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/LGA/Personal_Income'
    output_dir.mkdir(exist_ok=True)
    wb_obj.save(output_dir / path.name)